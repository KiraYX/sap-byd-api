from openpyxl import load_workbook

def write_df_to_excel(df, file_name, sheet_name, start_row=8, start_col=1):

    # Specify the path to your Excel file
    excel_file_path = f'../data/{file_name}'

    # Load the existing workbook
    book = load_workbook(excel_file_path)

    # Access the desired sheet
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        # Iterate over the DataFrame and write each value to the specified location
        for i in range(df.shape[0]):  # Loop through each row
            for j in range(df.shape[1]):  # Loop through each column
                # Write each cell in the DataFrame to the correct position in the Excel sheet
                sheet.cell(row=start_row + i, column=start_col + j, value=df.iat[i, j])

    # Save the workbook
    book.save(excel_file_path)
    print(f"DataFrame written to {excel_file_path} in sheet '{sheet_name}' starting from cell R{start_row}C{start_col} (without headers).")